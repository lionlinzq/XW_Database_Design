import json

from openpyxl import *


class ExcelOp(object):
    def __init__(self, file, sheet):
        self.file = file
        self.wb = load_workbook(self.file)
        sheets = self.wb.sheetnames
        if sheet:
            self.sheet = sheet
        else:
            sheet = sheets[0]
            self.sheet = sheets[0]
        self.ws = self.wb[self.sheet]

    # 获取表格的总行数和总列数
    def get_row_clo_num(self):
        rows = self.ws.max_row
        columns = self.ws.max_column
        return rows, columns

    # 获取某个单元格的值
    def get_cell_value(self, row, column):
        cell_value = self.ws.cell(row=row, column=column).value
        return cell_value

    # 获取某列的所有值
    def get_col_value(self, column):
        rows = self.ws.max_row
        column_data = []
        for i in range(1, rows + 1):
            cell_value = self.ws.cell(row=i, column=column).value
            column_data.append(cell_value)
        return column_data

    # 获取某行所有值
    def get_row_value(self, row):
        columns = self.ws.max_column
        row_data = []
        for i in range(1, columns + 1):
            cell_value = self.ws.cell(row=row, column=i).value
            row_data.append(cell_value)
        return row_data

    def get_all_rows_value(self):
        """

        :rtype: object
        """
        rows_data = []
        for i in range(self.ws.max_row):
            rows_data.append(self.get_row_value(i + 1))
        return rows_data

    # 设置某个单元格的值
    def set_cell_value(self, row, colunm, cellvalue):
        try:
            self.ws.cell(row=row, column=colunm).value = cellvalue
            self.wb.save(self.file)
        except:
            self.ws.cell(row=row, column=colunm).value = "writeFail"
            self.wb.save(self.file)

    def get_value(self):
        result = []
        for row in self.ws.rows:
            temp = []
            for cell in row:
                temp.append(cell.value)
            # print(temp)
            result.append(temp)
        return result


if __name__ == "__main__":
    # excel_op = ExcelOp(file="/Users/lin/PycharmProjects/pythonProject/xuanwu/data/报表任务及调度.xlsx")
    # print(excel_op.get_col_value(8))
    excel_op = ExcelOp("/Users/lin/PycharmProjects/pythonProject/xuanwu/data/智慧100 数据字典整理.xlsx")
    # rows, columns = excel_op.get_row_clo_num()
    # print(excel_op.get_row_value(1))
    # print("行数:", end='')
    # print(rows)
    # for i in range(rows):
    #     rowData = excel_op.get_row_value(i+1)
    #     if rowData[0] == 'wtn_bi_dim_product':
    #         print(i+1)
